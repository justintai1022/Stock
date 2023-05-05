#!/usr/bin/env python3
#########################################################################################################################################
#    File name: Stock.py                                                                                                                #
#    Author: Justin Tai                                                                                                                 #
#    Date created: 2022/02/02                                                                                                           #
#    Date last modified: 2023/01/24                                                                                                     #
#    Python Version: 3.10.0                                                                                                             #
#    Version: 4.1                                                                                                                       #
#########################################################################################################################################
#    Version 1.0 : 抓取前六年股價、股利並加已計算
#    Version 1.1 : 修改成判斷當年度的股利是否公佈，判斷完後再做近五年度的平均值計算
#    Version 2.0 : 新增成可重覆執行的功能並寫在一個excel上
#                  新增USB的log不顯現出來
#                  新增浮點數只取小數點後兩位
#    Version 2.1 : 新增上傳到google driver 可以抓取到股價的公式
#                  Sheet頁面增加股票名稱
#    Version 3.0 : 新增可以輸入股票名稱功能
#                  新增判斷輸入錯誤的股票代號和名稱
#    Version 3.1 : 修改只抓取前五年股價和股利
#    Version 3.2 : 要新增程式中斷後可以接續上一個sheet的作業
#    Version 3.3 : 將1.1的功能加回來，判斷當年度的股利是否公佈，判斷完後再做近五年度的平均值計算
#    Version 3.4 : 網頁程式碼有改，跟著修改
#    Version 4.0 : 新增判斷是否為年配的股票
#    Version 4.1 : 新增判斷為年配股票後，可以再次輸入股票代號。還有把進入股票頁面的功能改為函式
#########################################################################################################################################

from datetime import date
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
import os
from selenium.common.exceptions import NoSuchElementException as error
import re


#設定環境變數
PATH = Service('.\chromedriver.exe')
today = date.today()
d4 = today.strftime("%Y-%m-%d")
cash_dividend_tr = []
cash_dividend_td = []
cash_dividend = "cash_dividend"+"_"+d4+".xlsx"
years_tr = 4
title = ("年度","股利","年度最高股價","年度最低股價","股價平均","最低殖利率(%)","最高殖利率(%)","平均殖利率(%)")
stock_name_field = '/html/body/table[2]/tbody/tr/td[3]/table/tbody/tr[1]/td/table/tbody/tr/td[1]/table/tbody/tr[1]/th/table/tbody/tr/td[1]/nobr/a'
stock_HomePage = "https://goodinfo.tw/tw/index.asp"
wb = openpyxl.Workbook()
options = webdriver.ChromeOptions()

#Add argument change log to level 3 to avoid selenium bug about USB connect by Orsan
options.add_argument("–log-level=3")
options.add_argument("--no-sandbox")
options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome(service=PATH,options=options,service_log_path=os.devnull)

def enter_website(stock):
    #進入股票頁面
    stock_search.send_keys(stock)
    login = driver.find_element(By.XPATH , '//*[@id="frmStockSearch"]/input[2]').click()
    time.sleep(2)

# def stock_number(stock):
#     #判斷所輸入的是股票代號還是股票名稱，若是名稱則轉成股票代號，若是代號則返回代號 2022/02/04
#     if not bool(re.search(r'\d',stock)):
#         stock_name = (str(driver.find_element(By.XPATH,stock_name_field).text))
#         stock_no = stock_name.replace(stock,"")
#         return stock_no
#     else:
#         return stock

#程式開始
driver.get("https://goodinfo.tw/tw/index.asp")
#Get current dir
current_working_dir = os.getcwd()
path = (f'{current_working_dir}')
file =  path + '\\' + "stock_sheet_" + d4 + '.txt'

while True:
    #取得寫到第幾個sheet
    if os.path.exists(file):
        file =  path + '\\' + "stock_sheet_" + d4 + '.txt'
        f = open(file,'r')
        page = int(f.read())
        wb = openpyxl.load_workbook(cash_dividend)
    else:
        page = 0

    stock_search = driver.find_element(By.ID , "txtStockCode")
    stock = input ("請輸入股票名稱(代號)或按quit/q結束程式  \n") #2022/02/04新增可以輸入股票名稱
    if stock == "quit" or stock =="q" or stock=="Q":
         driver.close()
         break
    try:
        #進入股票頁面
        enter_website(stock)
        #stock = stock_number(stock)
        stock_name = (str(driver.find_element(By.XPATH,stock_name_field).text)) #抓取代號+名稱做為sheet的名字
        
        #進入股票基本資料頁面並判斷是否為年配的股票和可以再次輸入股票代號 2023/01/24
        while True:
            cash_BasicInfo_url = "https://goodinfo.tw/tw/BasicInfo.asp?STOCK_ID="+stock
            driver.get(cash_BasicInfo_url)
            time.sleep(2)
            cash_Policy = (str(driver.find_element(By.XPATH,'/html/body/table[2]/tbody/tr/td[3]/table[2]/tbody/tr[13]/td[4]/nobr').text))
            if cash_Policy != '每年':
               print ("該股票非年配股票:"+stock_name+"\n")
               driver.get(stock_HomePage)
               time.sleep(2)
               stock_search = driver.find_element(By.ID , "txtStockCode")
               stock = input ("請輸入股票名稱(代號)或按quit/q結束程式  \n")
               if stock == "quit" or stock =="q" or stock=="Q":
                    driver.close()
                    os._exit(0)
               enter_website(stock) #進入股票頁面
               #stock = stock_number(stock)
               stock_name = (str(driver.find_element(By.XPATH,stock_name_field).text)) #抓取代號+名稱做為sheet的名字
            else:
                break


        #進入股利政策頁面
        cash_dividend_url= "https://goodinfo.tw/tw/StockDividendPolicy.asp?STOCK_ID="+stock
        driver.get(cash_dividend_url)
        time.sleep(2)

        print ("請等待資料寫入到excel ...... \n")

        #建立excel
        stock_name = (str(driver.find_element(By.XPATH,'/html/body/table[2]/tbody/tr/td[3]/table[1]/tbody/tr/td[1]/table/tbody/tr[1]/th/table/tbody/tr/td[1]/nobr/a').text))
        wb.create_sheet(stock_name+"現金股利",page) #2022/02/04，Sheet頁面增加股票名稱
        wb.active = int(page)
        sheet = wb.active
        sheet.append (title)
        wb.save(cash_dividend)
                        
        soup = BeautifulSoup(driver.page_source,"html.parser")
        tables = soup.find(id="tblDetail")

        #抓取最近六年年度股利、股價和現金殖利率
        for j in range (0,6):
            if years_tr <= 9:

                trs = tables.find_all("tr")[years_tr] #從當年度開始取一整行的資料，例如當年度為2022年那就是tr[4]到2017年tr[9]
                
                for tr in trs:
                    cash_dividend_tr.append(tr.getText())
                #抓取特定資料(股利、股價和現金殖利率)
                cash_dividend_td[0:4] = (cash_dividend_tr[1],cash_dividend_tr[7],cash_dividend_tr[27],cash_dividend_tr[29])
                cash_dividend_td[0] = str(cash_dividend_td[0])
                #判斷資料中是否有 - 
                if cash_dividend_td[1] != '-':
                    #將股利、股價和現金殖利率形態改為浮點數
                    for i in range(1,4):
                        cash_dividend_td[i] = float(cash_dividend_td[i])
                else:
                    #將股利、股價和現金殖利率形態改為浮點數
                    for i in range(2,4):
                        cash_dividend_td[i] =float(cash_dividend_td[i])
                #計算股價平均、還有殖利率(最高、最低和平均)
                if cash_dividend_td[1] != '-':
                    price_avge = (cash_dividend_td [2] + cash_dividend_td[3]) / 2
                    cash_dividend_td = cash_dividend_td + [price_avge]
                    for i in range(2,5):
                        cash_dividend_avge = (cash_dividend_td[1] / cash_dividend_td [i]) * 100
                        cash_dividend_td = cash_dividend_td + [cash_dividend_avge]

                sheet.append (cash_dividend_td)
                wb.save(cash_dividend)
                cash_dividend_tr = []
                cash_dividend_td = []
                years_tr = years_tr+1

        #讀取excel檔
        wb = openpyxl.load_workbook(cash_dividend)
        wb.active = page
        ws = wb.active
        ws ['A8'].value = '平均'
        ws ['A9'].value = '昂貴價'
        ws ['A10'].value = '便宜價'
        ws ['A11'].value = '合理價'
        ws ['A12'].value = '股價' #2022/02/04新增
        ws ['B12'].value = '=GOOGLEFINANCE("TPE:'+stock+'"'+',"price")' #2022/02/04 新增可以放到google上抓取股價的公式
        
        #判斷當年度股利是否已公佈，若公佈則拿當年度的資料去計算近五年的平均值，若未公佈則用前一年度的資料去計算近五年的平均值
        if ws['B2'].value != '-':
            ws.delete_rows(7)
            for col in range(2,9):
                char = get_column_letter(col)
                ws[char+'7'] = f'=AVERAGE({char +"2"}:{char+"6"})'
        else:
            ws.delete_rows(2)
            for col in range(2,9):
                char = get_column_letter(col)
                ws[char+'7'] = f'=AVERAGE({char +"2"}:{char+"6"})'

        #將股價平均、最低最高平均殖利率換算成取小數第兩位
        for col in range(5,9):
            char = get_column_letter(col)
            for row in range(2,7):
                price = ws[char+str(row)].value
                ws[char+str(row)].value = f'=ROUND({price},2)'
        
        #計算出昂貴價、合理價和便宜價並取到小數點後兩位
        for col1 in range(2,3):
            char1 = get_column_letter(col1)
            col_num = 8
            for col2 in range (6,9):
                char2 = get_column_letter(col2)
                ws[char1+str(col_num)] = f'=ROUND({char1 +"7"} / {char2+"7"}*100,2)'
                col_num = int(col_num+1)

        wb.save(cash_dividend)               
        page = page+1
        years_tr = 4

        #記錄正要寫入第幾個sheet
        f = open (file,'w')
        print (page, file=f)
        f.close
        
        #回到首頁
        driver.get(stock_HomePage)

    #處理輸入錯誤的股票名稱和代號 2022/02/04
    except error:
        print ("找不到您所輸入的股票\n")